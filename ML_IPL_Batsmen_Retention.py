import pandas as pd
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score,precision_score,recall_score,f1_score,confusion_matrix
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import xlsxwriter  
import openpyxl
                                          # DATA - PREPARATION - STARTS
                                                     #Getting Input Files
print(" ")
print()
print("          Enter all file names with extension")
flename = input("Enter the file name to get Batsman's IPL Ranking : ")
workbook = load_workbook(filename=flename)
flename1 = input("Enter the file name to get Batsman's Orange-cap list : ")
workbook1 = load_workbook(filename=flename1)
flename2 = input("Enter the file name to get Last Year IPL Team Runs : ")
workbook2 = load_workbook(filename=flename2)
flename3 = input("Enter the file name to get Last Year's Individual IPL Runs by every Batsman : ")
workbook3 = load_workbook(filename=flename3)
workbook4 = xlsxwriter.Workbook('C:/Users/sathya narayanan/Documents/Data_Mining_Project/Result.xlsx')
workbook4a = load_workbook(filename="Result.xlsx")
                                                      #Getting Input Files

                                               #initializing needed variables
teams = []
runs = []
runs_at_win = []
sums =[]
dic = {}
dic2 = {}
sheet = workbook.active
sheet1 = workbook1.active
sheet2 = workbook2.active
sheet3 = workbook3.active
sheet4 = workbook4a.active
fname = 'C:/Users/sathya narayanan/Documents/Data_Mining_Project/Result.xlsx'
sheet["G1"]="RANK"
sheet["H1"]="POINTS"
sheet1["D1"]="POINTS"
sheet3["D1"]="POINTS AT OVERALL"
sheet3["E1"]="POINTS AT WINNING"
sheet3["F1"]="RECENT-FORM POINTS"
sheet4["C1"]="CONSISTENCY POINTS"
sheet4["D1"]="IPL RANK POINTS"
sheet4["E1"]="TOTAL POINTS"
sheet4["G1"]="RETENTION CHANCES"
                                                 #initializing needed variables

                                                      #Data pre-processing
for i in range(2,sheet.max_row+1):
   sheet.cell(row=i,column=7).value = i-1
for j in range(2,sheet.max_row+1):
   if sheet.cell(row=j,column=7).value <= 10:
       sheet.cell(row=j,column=8).value = 5
   elif sheet.cell(row=j,column=7).value <= 25:
       sheet.cell(row=j,column=8).value = 4
   elif sheet.cell(row=j,column=7).value <= 50:
       sheet.cell(row=j,column=8).value = 3
   elif sheet.cell(row=j,column=7).value <= 75:
       sheet.cell(row=j,column=8).value = 2
   elif sheet.cell(row=j,column=7).value <= 100:
       sheet.cell(row=j,column=8).value = 1
for i1 in range(2,sheet1.max_row+1):
   a = sheet1.cell(row=i1,column=1).value
   if a == 1 or a==2 or a==3 or a==4 or a==5:
       sheet1.cell(row=i1,column=4).value = 10
   if a == 6 or a == 7 or a==8 or a==9 or a==10:
       sheet1.cell(row=i1,column=4).value = 5
for k in range(2,sheet2.max_row+1):
      teams.append(sheet2.cell(row=k,column=1).value) 
      runs.append(sheet2.cell(row=k,column=2).value)
      runs_at_win.append(sheet2.cell(row=k,column=3).value)
for k in teams:
   for v in runs:
      dic[k] = v
      runs.remove(v)
      break
for k in teams:
   for v in runs_at_win:
      dic2[k] = v
      runs_at_win.remove(v)
      break
for y in range(2,sheet3.max_row+1):
  ans = (sheet3.cell(row=y,column=2).value * 50) / dic[sheet3.cell(row=y,column=3).value]
  ans2 = (sheet3.cell(row=y,column=2).value * 100) / dic2[sheet3.cell(row=y,column=3).value]
  sheet3.cell(row=y,column=4).value = ans
  sheet3.cell(row=y,column=5).value = ans2
  sheet3.cell(row=y,column=6).value = sheet3.cell(row=y,column=4).value + sheet3.cell(row=y,column=5).value
workbook.save(filename="Batterlist.xlsx")
workbook1.save(filename="orange_cap_list.xlsx")
workbook3.save(filename="Individual_Scores.xlsx")
                                                       #Data pre-processing

                                                         #Data Preparation
for i in range(1,sheet4.max_row+1):
   sheet4.cell(row=i,column=1).value = sheet3.cell(row=i,column=1).value
   sheet4.cell(row=i,column=2).value = sheet3.cell(row=i,column=6).value
for i in range(2,sheet4.max_row+1): 
  sheet4.cell(row=i,column=3).value = 0
  sheet4.cell(row=i,column=4).value = 0
for i in range(2,sheet4.max_row+1):
  for j in range(2,sheet4.max_row+1):
    if sheet1.cell(row=i,column=2).value == sheet4.cell(row=j,column=1).value:
        sheet4.cell(row=j,column=3).value = sheet4.cell(row=j,column=3).value + sheet1.cell(row=i,column=4).value 
for i in range(2,sheet.max_row+1):
  for j in range(2,sheet4.max_row+1):
    if sheet.cell(row=i,column=1).value == sheet4.cell(row=j,column=1).value:
        sheet4.cell(row=j,column=4).value = sheet.cell(row=i,column=8).value
for i in range(2,sheet4.max_row+1):
   sheet4.cell(row=i,column=5).value = sheet4.cell(row=i,column=4).value + sheet4.cell(row=i,column=3).value + sheet4.cell(row=i,column=2).value



                                                     #""" Regression Training Model """
for i in range(2,sheet4.max_row+1):
  if float(str(sheet4.cell(row=i,column=2).value)) >= float(30) or int(str(sheet4.cell(row=i,column=3).value)) >= 20 or int(str(sheet4.cell(row=i,column=4).value)) >=4:
       sheet4.cell(row=i,column=7).value = "       1"
  else:                                                             
       sheet4.cell(row=i,column=7).value = "       0"

                                                    # """ Regression Training Model """
                                                      #Saving output file
workbook4a.save(filename="Result.xlsx")
                                                             #Saving output file
                                                          #Data Preparation
                                                   # DATA - PREPARATION -ENDS
             
                                                 # MACHINE - LEARNING - MODEL
result = pd.read_csv('C:/Users/sathya narayanan/Downloads/result.csv')
X = result[['RECENT-FORM POINTS','CONSISTENCY POINTS','IPL RANK POINTS','TOTAL POINTS']].values
y = result['RETENTION CHANCES'].values
X_train,X_test,y_train,y_test = train_test_split(X,y,random_state = 30)
model = LogisticRegression()
model.fit(X_train,y_train)
m = model.fit(X_train,y_train)
y_pred = model.predict(X_test)  
print("Accuracy_Score : ",accuracy_score(y_test,y_pred))
print("Precision_Score : ",precision_score(y_test,y_pred))
print("Recall_Score : ",recall_score(y_test,y_pred))
print("F1_Score : ",f1_score(y_test,y_pred))
print("Confusion_Matrix : \n",confusion_matrix(y_test,y_pred)) # Matrix-format :tn,fp,fn,tp
if model.predict([['27.9438014521585','0','2','29.9438014521585']]) < 0.5:
    print("\n Player not Retained")
else:
    print("\n Player Retained")
plt.scatter(X_test[:,0],y_pred)       # To plot test-data
plt.show()